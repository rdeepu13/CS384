'''
Deepika Rajwar
1901EE19
'''
import os
import csv
import openpyxl

def generate_marksheet():
	path=".\output"
	roll_list=[] 	#stores all roll no
	with open("names-roll.csv",'r') as f1:
		reader=csv.reader(f1)
		Dict_from_f1={row[0]:row[1] for row in reader}
	f1.close()
	with open("grades.csv") as f:
		csv_file=csv.DictReader(f)
		for row in csv_file:
			roll=row["Roll"] 
			roll_list.append(roll)
			curr_sem="Sem"+row["Sem"]
			sub_code=row["SubCode"]
			filename=roll+".xlsx"
			filedir=os.path.join(path,filename)
			try:
				wb=openpyxl.load_workbook(filedir)
			except:
				wb=openpyxl.Workbook()
			try:
				sheet=wb[curr_sem]		
			except:
				sheet=wb.create_sheet(curr_sem)
				sheet.append(["S.No","Subject No","Subject Name","L-T-P","Credit","Subject Type","Grade"])
			list=[]
			with open("subjects_master.csv") as f2:
				r=csv.reader(f2)
				for i in r:
					if i[0]==sub_code:
						list=i
			f2.close()
			l=[1]+list+[row["Sub_Type"]]+[row["Grade"]]
			sheet.append(l)	
			wb.save(filedir)
	f.close() 
	Grading={'AA':10,'AB':9,'BB':8,'BC':7,'CC':6,'CD':5,'DD':4,'DD*':4,'F*':0,'F':0,'I':0}
	roll_list=set(roll_list)		#using set to remove duplicated rollno from list
	for filename in roll_list:
		file_name=filename+".xlsx"
		filedir=os.path.join(path,file_name)
		try:
			wb=openpyxl.load_workbook(filedir)
		except:
			continue
		sheet=wb.active
		sheet['A1']="Roll No"
		sheet['B1']=filename
		sheet['A2']="Name of Student"
		sheet['B2']=Dict_from_f1[filename]
		sheet['A3']="Discipline"
		st=filename.strip()
		br=st[4:6]
		sheet['B3']=br
		sheet['A4']="Semester No"
		c=1
		for i in range(2,10):
			sheet.cell(row=4,column=i).value=c;
			c+=1
		sheet['A5']="Semester wise Credit Taken"
		sheet['A6']="Spi"
		sheet['A7']="Total Credits Taken"
		sheet['A8']="Cpi"
		sheet.title="Overall"
		#Fixing the S.No in semester sheets
		credit=[]	#A5
		total_credit=[]	#A7
		sum_credit=0
		total_grade=0
		cpi=[]		#A8
		spi=[]		#A6
		for sheet in wb.worksheets:
			if sheet.title=='Overall':
				continue;
			c=1
			credit_sem=0
			grade_sem=0
			mr=sheet.max_row
			for i in range(2,mr+1):
				sheet.cell(row=i,column=1).value=c		#Fixing Serial No here
				credit_sem+=int(sheet.cell(row=i,column=5).value)
				grade_sem+=(int(Grading[sheet.cell(row=i,column=7).value]))*(int(sheet.cell(row=i,column=5).value))
				c+=1
			credit.append(credit_sem)
			spi.append(float("{:.2f}".format(grade_sem/credit_sem)))
			sum_credit+=credit_sem
			total_credit.append(sum_credit)
			total_grade+=grade_sem
			cpi.append(float("{:.2f}".format(total_grade/sum_credit)))
		sheet=wb['Overall']
		j=0;
		for i in range(2,10):
			sheet.cell(row=5,column=i).value=credit[j];		
			sheet.cell(row=6,column=i).value=spi[j];
			sheet.cell(row=7,column=i).value=total_credit[j];
			sheet.cell(row=8,column=i).value=cpi[j];
			j+=1
		wb.save(filedir)
	return

generate_marksheet()


