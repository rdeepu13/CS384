import csv
import openpyxl

wb=Workbook()

def output_by_subject():
    with open('./regtable_old.csv', 'r') as file:      # check and edit the path
	reader=csv.reader(file)
        sub = []
        for line in reader:
            row = line.split(',')
            if row[3] not in sub:
                sub.append(row[3])
        for i in range(0, len(sub)):
	    wb.save("{}.xlsx",sub[i])
	    sh=wb.create_sheet(sub[i])
            wb.load_workbook("{}.xlsx",sub[i])
	    workbook.write("rollno,register_sem,subno,sub_type\n")
	    wb.save("{}.xlsx",sub[i])
    with open('./regtable_old.csv', 'r') as reg:      # check and edit the path
	row=csv.DictReader(reg)
	del row[4:8]
	row.pop(2)
	if str(row[2]) in sub:
	    wb.load_workbook("{}.xlsx",format(row[2]))
            sh.append(row)
	    wb.save("{}.xlsx",format(row[2]))
    return

def output_individual_roll():
    with open("./regtable_old.csv", "r") as file:      # check and edit the path
        reader=csv.reader(file)
        roll = []
        for line in reader:
            if line[0] not in roll:
			roll.append(line[0])
        for i in range(0, len(roll)):
	    sheet=wb.active()
	    for k in range(1,5):
		sheet.cell(row=1,column=k).value=head[k-1]
	    wb.save("./output_individual_roll/{}.xlsx".format(roll[i]))
    with open("./regtable_old.csv", "r") as reg:      # check and edit the path
	m=csv.DictReader(reg)
	del row[4:8]
            row.pop(2)
        for i in range(0,len(roll)):
	    wb=load_workbook("./output_individual_roll/{}.xlsx".format(roll[i]))
	    sh=append(row)
	    wb.save("./output_individual_roll/{}.xlsx".format(roll[i]))
    return

output_individual_roll()
output_by_subject()