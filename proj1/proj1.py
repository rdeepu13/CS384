import os
import tkinter as tkin
from tkinter import filedialog
import csv
import pandas as pd
import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Font, Border
from openpyxl.styles.borders import Side
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

pos = float
neg = float

def err_msg(n):         # Shows the appropiate error message, if encountered
    err = tkin.Tk()
    err.title('Error Message')
    err.geometry('600x200')
    err_tup = ('No file selected!', 'No roll number with ANSWER is present, Cannot process!',
               'No marks entered, Invalid marking scheme!',
               'Marks entered is neither integer nor float, Invalid marking scheme!',
               'Positive marks entered for wrong answers!\nValid input is a negative marks or zero.',
               'Zero or negative marks entered for correct answers!\nValid input is a positive value.',
               'Unable to login!\nUsername or Password not entered.',
               'Unable to login!\nUsername is not a GMAIL ID.',
               'Unable to login and proceed!\nIncorrect Username and Password.')
    tkin.Label(err, text = '{}'.format(err_tup[n]), font = ('Arial', 10)).pack(padx = '25', pady = '20')
    tkin.Label(err, text = "Click 'Close' to exit program", font = ('Arial', 10)).pack(padx = '25', pady = '10')
    tkin.Button(err, text = 'Close', width = 15, command = exit).pack(padx = '50')
    err.mainloop()

def userinput():        # Processes the marking scheme input
    global pos, neg
    pos_g = pm.get()
    neg_g = nm.get()
    mark.destroy()
    if (pos_g == '' or neg_g == ''):
        err_msg(2)
    try:
        pos = float(pos_g)
        neg = float(neg_g)
        if (neg == 0):
            neg = float(0)
        if (neg > 0):
            err_msg(4)
        if (pos <= 0):
            err_msg(5)
    except ValueError:
        err_msg(3)

gui_mas = tkin.Tk()            # File explorer for 'master_roll.csv'
gui_mas.title('Browse for file')
gui_mas.geometry('300x90')
tkin.Label(gui_mas, text = "Select the file named 'master_roll.csv'.", font = ('Arial', 10)).pack(padx = '25', pady = '20')
mas_file = filedialog.askopenfilename(initialdir = '/', title = 'Browse for master_roll.csv', filetypes = [('CSV Files','*.csv')])
gui_mas.destroy()
if mas_file == '':
    err_msg(0)

gui_res = tkin.Tk()            # File explorer for 'responses.csv'
gui_res.title('Browse for file')
gui_res.geometry('300x90')
tkin.Label(gui_res, text = "Select the file named 'responses.csv'.", font = ('Arial', 10)).pack(padx = '25', pady = '20')
res_file = filedialog.askopenfilename(initialdir = '/', title = 'Browse for responses.csv', filetypes = [('CSV Files','*.csv')])
gui_res.destroy()
if res_file == '':
    err_msg(0)

mark = tkin.Tk()         # Window to input marking scheme 
pm = tkin.StringVar()
nm = tkin.StringVar()
mark.title('Marking scheme to be used')
tkin.Label(mark, text = "Marks for correct answer: ", font = ('Arial', 10)).grid(row = 0, column = 0, padx = '25', pady = '10')
tkin.Label(mark, text = "Marks for wrong answer: ", font = ('Arial', 10)).grid(row = 1, column = 0, padx = '25', pady = '10')
tkin.Button(mark, text = 'Close', width = 10, command = exit).grid(row = 2, column = 1, padx = '50', pady = '20')
tkin.Button(mark, text = 'Submit', width = 10, command = userinput).grid(row = 2, padx = '50', pady = '20')
pos_mark = tkin.Entry(mark, textvariable = pm)
pos_mark.grid(row = 0, column = 1, columnspan = 3, padx = '25')
neg_mark = tkin.Entry(mark, textvariable = nm)
neg_mark.grid(row = 1, column = 1, columnspan = 3, padx = '25')

res_df = pd.read_csv(res_file)       # DataFrame for the file 'responses.csv'
ans = res_df.loc[(res_df['Roll Number'].str.upper() == 'ANSWER')]
if ans.empty:
    err_msg(1)
res_df.fillna(value = '', inplace = True)
res_df.insert(loc = 6, column = 'Score_After_Negative', value = ['' for i in range(0, res_df.shape[0])])
res_df['statusAns'] = ''
corr_ans = tuple(res_df.loc[res_df['Roll Number'].str.upper() == 'ANSWER'].values[0][8:-1])         # Tuple of correct answers

def generate():         # Generates roll number wise marksheets
    gen.destroy()
    try:
        route = os.path.join("./marksheets/")
        os.mkdir(route)
    except:       # If folder 'marksheets' already exists
        done = tkin.Tk()
        done.title('Folder exists')
        done.geometry('500x240')
        tkin.Label(done, text = "Folder 'marksheets' already exists.\nSaving new marksheets in the existing folder.", font = ('Arial', 10)).pack(padx = '25', pady = '20')
        tkin.Label(done, text = "Click 'OK' to close this window and proceed.\nClick 'Close' to end program", font = ('Arial', 9)).pack(padx = '25', pady = '10')
        tkin.Button(done, text = 'OK', width = 10, command = done.destroy).pack(padx = '70', pady = '20', side = 'left')
        tkin.Button(done, text = 'Close', width = 10, command = exit).pack(padx = '70', pady = '20', side = 'right')
        done.mainloop()
    q_no = len(corr_ans)
    full_mark = float(q_no * pos)
    red_f = Font(size = 12, color = 'FF0000', name = 'Century')
    green_f = Font(size = 12, color = '008000', name = 'Century')
    blue_f = Font(size = 12, color = '0000FF', name = 'Century')
    bold_f = Font(size = 12, name = 'Century', bold = True)
    norm_f = Font(size = 12, name = 'Century')
    center_a = Alignment(horizontal = 'center')
    right_a = Alignment(horizontal = 'right')
    left_a = Alignment(horizontal = 'left')
    thin_bord = Border(left = Side(border_style = 'thin'),
                       right = Side(border_style = 'thin'),
                       top = Side(border_style = 'thin'),
                       bottom = Side(border_style = 'thin'))
    tab_head = ['', 'Right', 'Wrong', 'Not Attempt', 'Max']
    tab_in = ['No.', 'Marking', 'Total']
    with open(mas_file) as master:
        branch = csv.DictReader(master)
        for tree in branch:
            c_flag, w_flag, e_flag, dex, t_mark = 0, 0, 0 ,0, 0.00
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'quiz'
            logo = Image('./proj_1_logo.jpg')
            logo.height = 83.15
            logo.width = 642.52
            ws.add_image(logo, 'A1')
            ws.merge_cells('A5:E5')
            ws['A5'] = 'Mark Sheet'
            ws.cell(row = 5, column = 1).alignment = center_a
            ws.cell(row = 5, column = 1).font = Font(bold = True, underline = 'single', name = 'Century', size = 18)
            for j in ['A', 'B', 'C', 'D', 'E']:
                ws.column_dimensions[j].width = 17.72
            ws['A6'] = 'Name: '
            ws.cell(row = 6, column = 1).alignment = right_a
            ws.cell(row = 6, column = 1).font = norm_f
            ws['A7'] = 'Roll Number: '
            ws.cell(row = 7, column = 1).alignment = right_a
            ws.cell(row = 7, column = 1).font = norm_f
            ws['D6'] = 'Exam: '
            ws.cell(row = 6, column = 4).alignment = right_a
            ws.cell(row = 6, column = 4).font = norm_f
            ws['B6'] = tree['name']
            ws.cell(row = 6, column = 2).alignment = left_a
            ws.cell(row = 6, column = 2).font = bold_f
            ws['B7'] = tree['roll']
            ws.cell(row = 7, column = 2).alignment = left_a
            ws.cell(row = 7, column = 2).font = bold_f
            ws['E6'] = 'quiz'
            ws.cell(row = 6, column = 5).alignment = left_a
            ws.cell(row = 6, column = 5).font = bold_f
            a = 0
            for row in ws.iter_rows(min_row = 9, max_row = 9, min_col = 1, max_col = 5):
                for cell in row:
                    cell.value = tab_head[a]
                    a += 1
                    cell.font = bold_f
                    cell.alignment = center_a
            a = 0
            for row in ws.iter_cols(min_row = 10, max_row = 12, min_col = 1, max_col = 1):
                for cell in row:
                    cell.value = tab_in[a]
                    a += 1
                    cell.font = bold_f
                    cell.alignment = center_a
            ws['A15'] = 'Student Ans'
            ws['A15'].alignment = center_a
            ws['A15'].font = bold_f
            ws['B15'] = 'Correct Ans'
            ws['B15'].alignment = center_a
            ws['B15'].font = bold_f
            a = 0
            for row in ws.iter_cols(min_row = 16, max_row = (16 + q_no - 1), min_col = 2, max_col = 2):
                for cell in row:
                    cell.value = corr_ans[a]
                    a += 1
                    cell.font = blue_f
                    cell.alignment = center_a
            ws['B11'] = pos
            ws['B11'].alignment = center_a
            ws['B11'].font = green_f
            ws['C11'] = neg
            ws['C11'].alignment = center_a
            ws['C11'].font = red_f
            ws['D11'] = '0'
            ws['D11'].alignment = center_a
            ws['D11'].font = norm_f
            ws['E10'] = q_no
            ws['E10'].alignment = center_a
            ws['E10'].font = norm_f
            chk_df = res_df.loc[(res_df['Roll Number'].str.upper()) == (tree['roll'].upper())]
            if (chk_df.empty):
                res_df.at[len(res_df), 'Name'] = tree['name']
                res_df.at[len(res_df)-1, 'Roll Number'] = tree['roll']
                res_df.at[len(res_df)-1, 'Score_After_Negative'] = 'Absent'
                ws.delete_rows(8, (ws.max_row-7))
            else:
                dstu_ans = tuple(res_df.loc[(res_df['Roll Number'].str.upper()) == (tree['roll'].upper())].values[0][8:-1])
                dex = int(list(chk_df.index)[0])
                for a in range(0, q_no):
                    if (dstu_ans[a] == ''):
                        e_flag += 1
                        ws.cell(row = (16+a), column = 1).value = ""
                    else:
                        if (dstu_ans[a] == corr_ans[a]):
                            c_flag += 1
                            ws.cell(row = (16+a), column = 1).value = dstu_ans[a]
                            ws.cell(row = (16+a), column = 1).font = green_f
                            ws.cell(row = (16+a), column = 1).alignment = center_a
                        else:
                            w_flag += 1
                            ws.cell(row = (16+a), column = 1).value = dstu_ans[a]
                            ws.cell(row = (16+a), column = 1).font = red_f
                            ws.cell(row = (16+a), column = 1).alignment = center_a
                t_mark = float(((c_flag * pos) + (w_flag * neg)))
                if t_mark.is_integer():
                    t_mark = int(t_mark)
                if str(full_mark)[-2:] == ".0":
                    full_mark = int(full_mark)
                res_df.at[dex, 'Score_After_Negative'] = '{0} / {1}'.format(t_mark, full_mark)
                res_df.at[dex, 'statusAns'] = '[{0}, {1}, {2}]'.format(c_flag, w_flag, e_flag)
                ws['B10'] = c_flag
                ws['B10'].alignment = center_a
                ws['B10'].font = green_f
                ws['C10'] = w_flag
                ws['C10'].alignment = center_a
                ws['C10'].font = red_f
                ws['D10'] = e_flag
                ws['D10'].alignment = center_a
                ws['D10'].font = norm_f
                ws['B12'] = (c_flag * pos)
                ws['B12'].alignment = center_a
                ws['B12'].font = green_f
                ws['C12'] = (w_flag * neg)
                ws['C12'].alignment = center_a
                ws['C12'].font = red_f
                ws['E12'] = '{}/{}'.format(t_mark, full_mark)
                ws['E12'].alignment = center_a
                ws['E12'].font = blue_f
                for row in ws.iter_cols(min_row = 9, max_row = 12, min_col = 1, max_col = 5):
                    for cell in row:
                        cell.border = thin_bord
                for row in ws.iter_cols(min_row = 15, max_row = ws.max_row, min_col = 1, max_col = 2):
                    for cell in row:
                        cell.border = thin_bord
            wb.save('./marksheets/{}.xlsx'.format(tree['roll'].upper()))
    res_df.fillna(value = "", inplace = True)
    return

gui_mas.mainloop()
gui_res.mainloop()
mark.mainloop()

gen = tkin.Tk()        # Window for generation of roll number wise marksheets
gen.title('Generate Marksheets')
gen.geometry('600x200')
tkin.Label(gen, text = "Click 'Generate' to generate roll number wise marksheets.", font = ('Arial', 10)).pack(padx = '25', pady = '20')
tkin.Label(gen, text = "Click 'Close' to exit program.", font = ('Arial', 9)).pack(padx = '25', pady = '10')
tkin.Button(gen, text = "Generate", width = 15, command = generate).pack(padx = '70', pady = '10', side = 'left')
tkin.Button(gen, text = 'Close', width = 10, command = exit).pack(padx = '70', pady = '10', side = 'right')
gen.mainloop()

def generate_sheet():       # Generates the file 'concise_marksheet.csv'
    gen_sh.destroy()
    res_df.to_csv('./marksheets/concise_marksheet.csv', index = False)
    return

gen_sh = tkin.Tk()         # Window for generation of 'concise_marksheet.csv'
gen_sh.title('Generate Concise Marksheet')
gen_sh.geometry('600x280')
tkin.Label(gen_sh, text = "Roll number wise marksheets successfully generated.", font = ('Arial', 14)).pack(padx = '25', pady = '30')
tkin.Label(gen_sh, text = "Click 'Generate' to concise marksheet.", font = ('Arial', 10)).pack(padx = '25', pady = '5')
tkin.Label(gen_sh, text = "Click 'Close' to exit program.", font = ('Arial', 9)).pack(padx = '25', pady = '10')
tkin.Button(gen_sh, text = "Generate", width = 15, command = generate_sheet).pack(padx = '70', pady = '10', side = 'left')
tkin.Button(gen_sh, text = 'Close', width = 10, command = exit).pack(padx = '70', pady = '10', side = 'right')
gen_sh.mainloop()

mail_df = res_df[['Roll Number', 'Email address', 'IITP webmail']]

def get_det():         # Process the login details input
    mail.destroy()
    id_in = id.get()
    pass_in = pass_id.get()
    if (id_in == '' or pass_in == ''):
        err_msg(6)
    if (id_in[-10:] != '@gmail.com'):
        err_msg(7)
    id_in = id_in.strip()
    pass_in = pass_in.strip()
    send_mail(id_in, pass_in)
    return

def send_mail(user_id, password):         # Sends mails
    try:
        serve = smtplib.SMTP('smtp.gmail.com', 587)
        serve.starttls()
        serve.login(user_id, password)
        with open(mas_file) as master:
            branch = csv.DictReader(master)
            for tree in branch:
                to_list = []
                chk_pd = mail_df.loc[(mail_df['Roll Number'].str.upper()) == (tree['roll'].upper())]
                if ((chk_pd['Email address'].empty) & (chk_pd['IITP webmail'].empty)):
                    pass
                elif ((chk_pd['Email address'].empty) & (~chk_pd['IITP webmail'].empty)):
                    to_list.append(str(chk_pd['IITP webmail'].values[0]).strip())
                elif ((~chk_pd['Email address'].empty) & (chk_pd['IITP webmail'].empty)):
                    to_list.append(str(chk_pd['Email address'].values[0]).strip())
                else:
                    to_list.append(str(chk_pd['Email address'].values[0]).strip())
                    to_list.append(str(chk_pd['IITP webmail'].values[0]).strip())                              
                mail_bod = MIMEMultipart()
                mail_bod['Subject'] = "Quiz Marksheet"
                mail_bod['From'] = user_id
                mail_bod['To'] = ", ".join(to_list)
                mail_text = "Quiz exam marks are attached for reference.\n+{0} for correct, {1} for wrong.\n--".format(str(pos), str(neg))
                mail_bod.attach(MIMEText(mail_text))
                attach_f = MIMEBase('application', "octet-stream")
                attach_f.set_payload(open("./marksheets/{}.xlsx".format(tree['roll'].upper()), "rb").read())
                encoders.encode_base64(attach_f)
                attach_f.add_header('Content-Disposition', 'attachment; filename="{}.xlsx"'.format(tree['roll'].upper()))
                mail_bod.attach(attach_f)
                serve.sendmail(user_id, to_list, mail_bod.as_string())
                to_list.clear()
        serve.quit()
    except smtplib.SMTPAuthenticationError:      # In case of incorrect login details
        err_msg(8)     
    return

mail_msg = tkin.Tk()         # Window to ask for sending mails
mail_msg.title('Send Mail')
mail_msg.geometry('500x280')
tkin.Label(mail_msg, text = "Concise marksheet successfully generated.", font = ('Arial', 14)).pack(padx = '25', pady = '30')
tkin.Label(mail_msg, text = "Click 'Send' to send mail with marksheets.", font = ('Arial', 10)).pack(padx = '25', pady = '5')
tkin.Label(mail_msg, text = "Click 'Close' to exit program.", font = ('Arial', 9)).pack(padx = '25', pady = '10')
tkin.Button(mail_msg, text = "Send", width = 10, command = mail_msg.destroy).pack(padx = '80', side = 'left')
tkin.Button(mail_msg, text = 'Close', width = 10, command = exit).pack(padx = '70', pady = '10', side = 'right')
mail_msg.mainloop()

mail = tkin.Tk()          # Window to input login details
id = tkin.StringVar()
pass_id = tkin.StringVar()
mail.title('Login details')
mail.geometry('500x200')
tkin.Label(mail, text = "Use GMAIL account details", font = ('Arial', 12)).grid(row = 0, column = 0, columnspan = 24, padx = '25', pady = '10')
tkin.Label(mail, text = "User ID: ", font = ('Arial', 10)).grid(row = 1, column = 0, padx = '25', pady = '10')
tkin.Label(mail, text = "Password: ", font = ('Arial', 10)).grid(row = 2, column = 0, padx = '25', pady = '10')
tkin.Button(mail, text = 'Close', width = 10, command = exit).grid(row = 3, column = 1, padx = '60', pady = '20')
tkin.Button(mail, text = 'Submit', width = 10, command = get_det).grid(row = 3, padx = '60', pady = '20')
user = tkin.Entry(mail, textvariable = id, width = 35)
user.grid(row = 1, column = 1, padx = '25')
pass_user = tkin.Entry(mail, textvariable = pass_id, show = "*", width = 35)
pass_user.grid(row = 2, column = 1, padx = '25')
mail.mainloop()

end_msg = tkin.Tk()         # Window to inform the end of program
end_msg.title('End of program')
end_msg.geometry('500x240')
tkin.Label(end_msg, text = "Mails successfully sent.", font = ('Arial', 14)).pack(padx = '25', pady = '30')
tkin.Label(end_msg, text = "Click 'Exit' to exit program", font = ('Arial', 9)).pack(padx = '25', pady = '5')
tkin.Button(end_msg, text = "Exit", width = 15, command = exit).pack(padx = '50', pady = '20')
end_msg.mainloop()